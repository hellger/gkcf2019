import requests
import re
from requests.exceptions import RequestException
from openpyxl import Workbook
from openpyxl import load_workbook

#定义一个方法，用来获取学生成绩存放的url,并访问这个url，然后返回网页内容
def get_url(url,user,pa,header):
    try:
        d = {'user':user,'pass':pa}
        r = requests.post(url, data=d,headers=header)
        rr=r.text.replace("\/","/")
        rrr=re.search('http.*html',rr)
        ru=requests.get(rrr.group())
        ru.encoding='utf-8'
        return ru.text
    except RequestException:
        return None

#h定义一个方法，用来解析网页内容
def parse_page(html):
    pattern = re.compile('<ul>.*?<li>考生姓名:<span>(.*?)\s+</span>.*?准考证号:<span>(\d+)</span>.*?文.*?class="bold">.*?(\d+)</em>.*?学.*?bold">.*?(\d+)</em>.*?外.*?语.*?bold">.*?(\d+).*?综合.*?bold">.*?(\d+).*?量化成绩.*?(\d+)</em>.*?总成绩.*?(\d+)\+?(\d+)?\(?(....)?\)?.*?次.*?(\d+)',re.S)
    items = re.findall(pattern,html)
    for i in items:
        for j in i:
            print (j, end = '\t')
    print('\r')
    return items

#定义一个方法，用来创建一个Excel文件保存成绩
def create_xlsx(name):
    wb = Workbook()
    ws = wb.active
    ws.title = '2019高考成绩'
    ws['A1'] = '姓名'
    ws['B1'] = '准考证号'
    ws['C1'] = '语文'
    ws['D1'] = '数学'
    ws['E1'] = '外语'
    ws['F1'] = '综合'
    ws['G1'] = '量化成绩'
    ws['H1'] = '总分'
    ws['I1'] = '加分'
    ws['J1'] = '加分类型'
    ws['K1'] = '全省排名'
    wb.save(filename = name)

#将成绩保存到之前创建的Excel中
def write_xlsx(items,m,name):
    wb = load_workbook(name)
    ws = wb.active
    for r in items:
        ws.cell(row=m,column=1).value = r[0]
        ws.cell(row=m,column=2).value = r[1]
        ws.cell(row=m,column=3).value = int(r[2])
        ws.cell(row=m,column=4).value = int(r[3])
        ws.cell(row=m,column=5).value = int(r[4])
        ws.cell(row=m,column=6).value = int(r[5])
        ws.cell(row=m,column=7).value = int(r[6])
        ws.cell(row=m,column=8).value = int(r[7])
        ws.cell(row=m,column=9).value = r[8]
        ws.cell(row=m,column=10).value = r[9]
        ws.cell(row=m,column=11).value = int(r[10])
    wb.save(filename = '2019高考成绩.xlsx')

#定义main方法，变量user表示开始的准考证号
def main():
    t='姓名\t准考证号\t语文\t数学\t外语\t综合\t量化\t总分\t加分\t加分类型\t全省排名'
    post_url = "http://www.ynzs.cn/2019gkcf/check.php?action=query"
    hander = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.87 Safari/537.36','Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8','Accept': 'application/json, text/javascript, */*; q=0.01','X-Requested-With': 'XMLHttpRequest','DNT': '1','Referer': 'http://www.ynzs.cn/2019gkcf/web.html','Accept-Encoding': 'gzip, deflate','Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8','Host': 'www.ynzs.cn'}
    create_xlsx('2019高考成绩.xlsx')
    wb = load_workbook('用户名和密码.xlsx',read_only = True)
    ws = wb.active
    row = 2
    column_username = 1
    column_password = 2
    print('★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★\n★★★★★★★★★★★★★★★★★★★★★★正在查询中★★★★★★★★★★★★★★\n★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★')
    print(t)
    while (row <= ws.max_row):
        username = ws.cell(row=row,column=column_username)
        password = ws.cell(row=row,column=column_password)
        html = get_url(post_url,username.value,password.value,hander)
        w=parse_page(html)
        write_xlsx(w,row,'2019高考成绩.xlsx')
        row =row+1
    print('查询完成，结果保存在当前目录下。\n注意：表格中如果有空行，说明这一行对应的账号或密码很可能不正确，请手动核实。')
if __name__=='__main__':
    main()